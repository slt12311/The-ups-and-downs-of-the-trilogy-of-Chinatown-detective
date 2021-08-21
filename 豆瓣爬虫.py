# 豆瓣爬虫
import re
from urllib.error import HTTPError
from bs4 import BeautifulSoup
import random
import time
from openpyxl import load_workbook


def coo_regular(cookie):
    coo = {}
    for k_v in cookie.split(';'):
        k, v = k_v.split('=', 1)
        coo[k.strip()] = v.replace('"', '')
    return coo

def getRequest(url):
    user_agents = [
        'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36 OPR/26.0.1656.60',
        'Opera/8.0 (Windows NT 5.1; U; en)',
        'Mozilla/5.0 (Windows NT 5.1; U; en; rv:1.8.1) Gecko/20061208 Firefox/2.0.0 Opera 9.50',
        'Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; en) Opera 9.50',
        'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:34.0) Gecko/20100101 Firefox/34.0',
        'Mozilla/5.0 (X11; U; Linux x86_64; zh-CN; rv:1.9.2.10) Gecko/20100922 Ubuntu/10.10 (maverick) Firefox/3.6.10',
        'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/534.57.2 (KHTML, like Gecko) Version/5.1.7 Safari/534.57.2 ',
        'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.71 Safari/537.36',
        'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.11 (KHTML, like Gecko) Chrome/23.0.1271.64 Safari/537.11',
        'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US) AppleWebKit/534.16 (KHTML, like Gecko) Chrome/10.0.648.133 Safari/534.16',
        'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/30.0.1599.101 Safari/537.36',
        'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; rv:11.0) like Gecko',
        'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.11 (KHTML, like Gecko) Chrome/20.0.1132.11 TaoBrowser/2.0 Safari/536.11',
        'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/21.0.1180.71 Safari/537.1 LBBROWSER',
        'Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; QQDownload 732; .NET4.0C; .NET4.0E)',
        'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.84 Safari/535.11 SE 2.X MetaSr 1.0',
        'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Trident/4.0; SV1; QQDownload 732; .NET4.0C; .NET4.0E; SE 2.X MetaSr 1.0) ',
        "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; AcooBrowser; .NET CLR 1.1.4322; .NET CLR 2.0.50727)",
        "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.0; Acoo Browser; SLCC1; .NET CLR 2.0.50727; Media Center PC 5.0; .NET CLR 3.0.04506)",
        "Mozilla/4.0 (compatible; MSIE 7.0; AOL 9.5; AOLBuild 4337.35; Windows NT 5.1; .NET CLR 1.1.4322; .NET CLR 2.0.50727)",
        "Mozilla/5.0 (Windows; U; MSIE 9.0; Windows NT 9.0; en-US)",
        "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Win64; x64; Trident/5.0; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 2.0.50727; Media Center PC 6.0)",
        "Mozilla/5.0 (compatible; MSIE 8.0; Windows NT 6.0; Trident/4.0; WOW64; Trident/4.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 1.0.3705; .NET CLR 1.1.4322)",
        "Mozilla/4.0 (compatible; MSIE 7.0b; Windows NT 5.2; .NET CLR 1.1.4322; .NET CLR 2.0.50727; InfoPath.2; .NET CLR 3.0.04506.30)",
        "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN) AppleWebKit/523.15 (KHTML, like Gecko, Safari/419.3) Arora/0.3 (Change: 287 c9dfb30)",
        "Mozilla/5.0 (X11; U; Linux; en-US) AppleWebKit/527+ (KHTML, like Gecko, Safari/419.3) Arora/0.6",
        "Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.8.1.2pre) Gecko/20070215 K-Ninja/2.1.1",
        "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9) Gecko/20080705 Firefox/3.0 Kapiko/3.0",
        "Mozilla/5.0 (X11; Linux i686; U;) Gecko/20070322 Kazehakase/0.4.5",
        "Mozilla/5.0 (X11; U; Linux i686; en-US; rv:1.9.0.8) Gecko Fedora/1.9.0.8-1.fc10 Kazehakase/0.5.6",
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.56 Safari/535.11",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_3) AppleWebKit/535.20 (KHTML, like Gecko) Chrome/19.0.1036.7 Safari/535.20",
        "Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; fr) Presto/2.9.168 Version/11.52"
    ]
    import requests
    cookie = ' ll="118090"; bid=LJi95RJ5m2I; __gads=ID=b213dfa8a10a562c-220a7b88a4ca0073:T=1627956247:RT=1627956247:S=ALNI_MbJrCzM4gQZi78PQeevpEXEkYWf4g; push_noty_num=0; push_doumail_num=0; douban-fav-remind=1; __yadk_uid=uWtXuihLf6cJt4uvEvLIpSfcG8gWhoEb; __utmc=30149280; _pk_ref.100001.8cb4=%5B%22%22%2C%22%22%2C1628038491%2C%22https%3A%2F%2Faccounts.douban.com%2Fpassport%2Flogin%3Fredir%3Dhttps%3A%2F%2Fwww.douban.com%2Fpeople%2F72834086%2F%22%5D; ct=y; __utmv=30149280.20253; _ga=GA1.2.112366083.1627976568; _gid=GA1.2.977337339.1628039213; __utma=30149280.112366083.1627976568.1628037741.1628040739.3; __utmz=30149280.1628040739.3.2.utmcsr=baidu|utmccn=(organic)|utmcmd=organic; _pk_id.100001.8cb4=57d1d379c4fc171c.1627956215.6.1628042430.1627978324.; __utmb=30149280.48.8.1628042429998; dbcl2="202535621:Ex5xUH5UYps"; ck=_Zwk'
    cookies = coo_regular(cookie)

    headers = {
        'User-Agent': str(random.choice(user_agents)),
        'Connection': 'keep-alive'
    }
    # 随机用一个请求头，开始访问地址
    try:
        req = requests.get(url=url,headers=headers, cookies=cookies)
    except HTTPError:
        pass
    # 返回结果
    return req

def getUser(url):
    req = getRequest(url).text
    # 读取数据(data得到所有数据)
    # 输出爬取到的所有数据，进制形式显示
    # 定义soup对象，解析网页
    soup = BeautifulSoup(req, "html.parser")
    if '页面不存在' in str(soup):
        return
    # 观影量
    num = str(soup.find('div', class_='info').find('h1'))
    num = num[num.find('(')+1 : num.find(')')]

    # 热点标签
    tag = soup.find_all('li', class_='clearfix')
    tag_resu = []
    for i in tag:
        tag0 = str(i.find('a'))
        pattern0 = re.compile(
            r'^<a .*>(.*)</a>.*')
        tag1 = pattern0.match(tag0).group(1)
        tag_resu.append(tag1)

    return num,tag_resu




# 封装函数，爬取数据
def getData(url,commentAll):
    # 获取处理后的请求
    req = getRequest(url).text
    soup = BeautifulSoup(req,"html.parser")
    # 找到装有所有评论的id名为comments的div
    # ["数据"]  数组里只有一个元素----数据
    comments = soup.select("#comments")[0]
    # print(comments)
    # 读取到每一条评论，div的class名为comment-item
    items = comments.select(".comment-item")
    # print(items)
    # 循环遍历每一条评论
    for i in items:
        # 找到装着用户名和星级的span标签，class名为comment-info
        good = str(i.select(".comment-vote")[0]).strip().replace("\n", "")
        pattern0 = re.compile(r'^<span class="comment-vote"><span class="votes vote-count">(.*)</span><input type="hidden".*')
        click = pattern0.match(good).group(1)

        timeof = str(i.select(".comment-time")[0]).strip().replace("\n", "")
        pattern0 = re.compile(
            r'^<span class="comment-time".*>(.*)</span>')
        timere = pattern0.match(timeof).group(1).strip()

        info = i.select(".comment-info")[0]
        resource = str(info).strip().replace("\n", "")
        pattern = re.compile(r'^<span class="comment-info"><a class="" href="(.*)">.*</a><span>')
        # 读出用户名的a标签里面的字符串用户名 [<a></a>]
        # author = info.select("a")[0].string  数据在列表里
        author = info.find("a").string
        path = (pattern.match(resource).group(1)+'collect').replace('www','movie')
        # print(author)
        # 取星级，找到装着星级的span标签，读取title值
        # ["看过"，星级，时间]
        star = info.select("span")[1]["title"]
        # 取评论，找到class名为short的p标签
        short = i.select(".short")[0].string
        # 将 用户名、星级、评论 装入在字典里面
        talk = {"author":author,"star":star,"short":short, "path":path, 'click':click, 'time':timere}
        # print(talk)
        # 将字典类型的数据，加到列表里面
        commentAll.append(talk)
    # 返回整个列表
    return commentAll

# 封装函数，把数据装入表格中
def writeInto(commentAll):
    tb_resource = load_workbook('doubanb.xlsx')
    tb_1 = tb_resource['唐探3']
    loc = tb_1.max_row
    for i, element in enumerate(commentAll):
        loc += 1
        # # 读取每一个字段  用户名、星级、评论
        # tb_1[''.join(['A', str(loc)])] = element['author']
        # tb_1[''.join(['D', str(loc)])] = element['time']
        # tb_1[''.join(['E', str(loc)])] = element['star']
        # tb_1[''.join(['F', str(loc)])] = element['short']
        #
        # result = getUser(element['path'])
        # print(result)
        # if(result):
        #     tb_1[''.join(['B', str(loc)])] = result[0]
        #     tb_1[''.join(['C', str(loc)])] = '、'.join(result[1][:5])
        # time.sleep(random.choice(range(3,10)))

        tb_1[''.join(['A', str(loc)])] = element['time']
        tb_1[''.join(['B', str(loc)])] = element['star']
        tb_1[''.join(['C', str(loc)])] = element['short']
        print(element)
        tb_resource.save('doubanb.xlsx')


# 函数的入口
# 直接输入main，有提示
if __name__ == '__main__':
    # 初始化一个空列表,将得到的所有数
    # range()产生序列 0.1.2,爬取3页
    for i in range(0,20):
        commentAll = []
        # 爬取的网页地址
        # limit=20 每一页读取20条数据
        # start = 80  从第几条读取数据 20-39  40-59  60-79 80-99
        url = "https://movie.douban.com/subject/27619748/comments?start=%d&limit=20&sort=new_score&status=P"%(i*20+162)
        # 调用函数，爬取数据
        getData(url,commentAll)
        time.sleep(3)
        writeInto(commentAll)
        # 每爬取一个页面数据，休息10秒，防止被封号


    # 将表格用 记事本 打开，另存为ANSI格式
    # 如果你要操作数据，还要转回utf-8
