import json
import random
import time
import requests
from openpyxl import load_workbook


class MaoYan(object):
    """
    猫眼评论爬虫，爬取电影《流浪地球》的评论和评分
    """

    def coo_regular(self,cookie):
        coo = {}
        for k_v in cookie.split(';'):
            k, v = k_v.split('=', 1)
            coo[k.strip()] = v.replace('"', '')
        return coo

    def __init__(self):
        """
        初始化函数
        :param
        headers: 请求头
        time: 当前时间戳
        premiere_time: 首映时间的时间戳
        """
        self.user_agents = [
            'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36 OPR/26.0.1656.60',
            'Opera/8.0 (Windows NT 5.1; U; en)',
            'Mozilla/5.0 (Windows NT 5.1; U; en; rv:1.8.1) Gecko/20061208 Firefox/2.0.0 Opera 9.50',
            'Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; en) Opera 9.50',
            'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:34.0) Gecko/20100101 Firefox/34.0',
            'Mozilla/5.0 (X11; U; Linux x86_64; zh-CN; rv:1.9.2.10) Gecko/20100922 Ubuntu/10.10 (maverick) Firefox/3.6.10',
            'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/534.57.2 (KHTML, like Gecko) Version/5.1.7 Safari/534.57.2 ',
            'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.71 Safari/537.36',
            'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.11 (KHTML, like Gecko) Chrome/23.0.1271.64 Safari/537.11',
            'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US) AppleWebKit/534.16 (KHTML, like Gecko) Chrome/10.0.648.133 Safari/534.16',
            'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/30.0.1599.101 Safari/537.36',
            'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; rv:11.0) like Gecko',
            'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.11 (KHTML, like Gecko) Chrome/20.0.1132.11 TaoBrowser/2.0 Safari/536.11',
            'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/21.0.1180.71 Safari/537.1 LBBROWSER',
            'Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; QQDownload 732; .NET4.0C; .NET4.0E)',
            'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.84 Safari/535.11 SE 2.X MetaSr 1.0',
            'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Trident/4.0; SV1; QQDownload 732; .NET4.0C; .NET4.0E; SE 2.X MetaSr 1.0) ',
            "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; AcooBrowser; .NET CLR 1.1.4322; .NET CLR 2.0.50727)",
            "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.0; Acoo Browser; SLCC1; .NET CLR 2.0.50727; Media Center PC 5.0; .NET CLR 3.0.04506)",
            "Mozilla/4.0 (compatible; MSIE 7.0; AOL 9.5; AOLBuild 4337.35; Windows NT 5.1; .NET CLR 1.1.4322; .NET CLR 2.0.50727)",
            "Mozilla/5.0 (Windows; U; MSIE 9.0; Windows NT 9.0; en-US)",
            "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Win64; x64; Trident/5.0; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 2.0.50727; Media Center PC 6.0)",
            "Mozilla/5.0 (compatible; MSIE 8.0; Windows NT 6.0; Trident/4.0; WOW64; Trident/4.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 1.0.3705; .NET CLR 1.1.4322)",
            "Mozilla/4.0 (compatible; MSIE 7.0b; Windows NT 5.2; .NET CLR 1.1.4322; .NET CLR 2.0.50727; InfoPath.2; .NET CLR 3.0.04506.30)",
            "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN) AppleWebKit/523.15 (KHTML, like Gecko, Safari/419.3) Arora/0.3 (Change: 287 c9dfb30)",
            "Mozilla/5.0 (X11; U; Linux; en-US) AppleWebKit/527+ (KHTML, like Gecko, Safari/419.3) Arora/0.6",
            "Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.8.1.2pre) Gecko/20070215 K-Ninja/2.1.1",
            "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9) Gecko/20080705 Firefox/3.0 Kapiko/3.0",
            "Mozilla/5.0 (X11; Linux i686; U;) Gecko/20070322 Kazehakase/0.4.5",
            "Mozilla/5.0 (X11; U; Linux i686; en-US; rv:1.9.0.8) Gecko Fedora/1.9.0.8-1.fc10 Kazehakase/0.5.6",
            "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.56 Safari/535.11",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_3) AppleWebKit/535.20 (KHTML, like Gecko) Chrome/19.0.1036.7 Safari/535.20",
            "Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; fr) Presto/2.9.168 Version/11.52"
        ]

        cookie = 'Cookie: _lxsdk_cuid=17b1bfe809cc8-045731578390ca-3e604809-e1000-17b1bfe809cc8; uuid_n_v=v1; iuuid=225605E0F71C11EB8219F16D87CE7293EEF48635AFE447178CB137CE42A06D8E; _lx_utm=utm_source%3DBaidu%26utm_medium%3Dorganic; ci=125%2C%E5%BC%A0%E5%AE%B6%E5%8F%A3; ci=125%2C%E5%BC%A0%E5%AE%B6%E5%8F%A3; ci=125%2C%E5%BC%A0%E5%AE%B6%E5%8F%A3; _lxsdk=974B4770F6C711EBA6613F74D8FD4754AEE87491AE394422ADAC32A2D0EE4444; Hm_lvt_703e94591e87be68cc8da0da7cbd0be2=1628299624,1628299636,1628301704,1628304067; __mta=20378247.1628262269866.1628304193402.1628304349672.32; Hm_lpvt_703e94591e87be68cc8da0da7cbd0be2=1628304351; _lxsdk_s=17b1df09e63-ab4-624-25c%7C%7C213'
        self.cookie = self.coo_regular(cookie)




    def get_comment(self):
        """
        爬取首映到当前时间的电影评论
        :param
        url: 评论真实请求的url，参数ts为时间戳
        :return: None
        """
        url = 'https://m.maoyan.com/mmdb/comments/movie/247300.json?_v_=yes&offset={}&startTime=2016-01-10%2010:08:01'
        for i in range(1,100):
            headers = {
                'User-Agent': str(random.choice(self.user_agents)),
                'Connection': 'keep-alive'
            }
            req_url = url.format(str(i * 15))
            res = requests.get(req_url, headers=headers, cookies=self.cookie).text
            print(req_url)
            time.sleep(random.choice(range(5,10)))
            comment_store = []
            count = 0
            for com in json.loads(res)['cmts']:
                comment = {'content': com['content'], 'gender': com['gender'] if 'gender' in com else '',
                           'cityName': com['cityName'] if 'cityName' in com else '', 'score': com['score'],
                           'time': com['time']}
                comment_store.append(comment)
                count += 1
                if count == 15:
                    self.time = com['time']

            self.writeInto(comment_store)

    def writeInto(self, commentAll):
        tb_resource = load_workbook('douban.xlsx')
        tb_1 = tb_resource['唐探11']
        loc = tb_1.max_row
        for i, element in enumerate(commentAll):
            loc += 1
            tb_1[''.join(['A', str(loc)])] = element['content']
            tb_1[''.join(['B', str(loc)])] = element['gender']
            tb_1[''.join(['C', str(loc)])] = element['cityName']
            tb_1[''.join(['D', str(loc)])] = element['score']
            tb_1[''.join(['E', str(loc)])] = element['time']
            print(element)
        tb_resource.save('douban.xlsx')


if __name__ == '__main__':
    my = MaoYan()
    my.get_comment()