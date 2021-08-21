import math

import xlwt

import openpyxl
from openpyxl.packaging import workbook


class proceed:


    def SentenceBreak(self):
        workbook = openpyxl.load_workbook("C:\\Users\\ysdong\\Desktop\\数学\\应用统计与R语言\\大作业\\doubanb.xlsx")
        sheets = workbook.get_sheet_names()
        booksheet = workbook.get_sheet_by_name(sheets[2])

        newsheet = workbook.get_sheet_by_name(sheets[5])
        cur=booksheet.cell(row=1,column=2).value
        id=booksheet.cell(row=1,column=1).value

        i=1
        newrow=1

        while cur!=None:

            while len(cur)!=0:
                firstSymbol=len(cur)

                tempPos=cur.find("！")
                if tempPos<firstSymbol and tempPos!=-1:
                    firstSymbol=tempPos

                tempPos=cur.find("，")
                if tempPos<firstSymbol and tempPos!=-1:
                    firstSymbol=tempPos

                tempPos=cur.find("。")
                if tempPos<firstSymbol and tempPos!=-1:
                    firstSymbol=tempPos

                tempPos=cur.find("!")
                if tempPos<firstSymbol and tempPos!=-1:
                    firstSymbol=tempPos

                tempPos=cur.find(",")
                if tempPos<firstSymbol and tempPos!=-1:
                    firstSymbol=tempPos

                tempPos=cur.find("?")
                if tempPos<firstSymbol and tempPos!=-1:
                    firstSymbol=tempPos

                tempPos=cur.find("？")
                if tempPos<firstSymbol and tempPos!=-1:
                    firstSymbol=tempPos

                tempPos = cur.find(" ")
                if tempPos < firstSymbol and tempPos != -1:
                    firstSymbol = tempPos

                tempPos = cur.find("\n")
                if tempPos < firstSymbol and tempPos != -1:
                    firstSymbol = tempPos


                newSen=cur[0:firstSymbol]
                print(newSen)
                cur=cur[firstSymbol+1:]
                newsheet.cell(newrow, 1, id)
                newsheet.cell(newrow,2,newSen)
                newrow=newrow+1


            i=i+1
            cur=booksheet.cell(row=i,column=2).value
            id = booksheet.cell(row=i, column=1).value


        workbook.save("C:\\Users\\ysdong\\Desktop\\数学\\应用统计与R语言\\大作业\\doubanb.xlsx")
        workbook.close()

    def factorCal(self):
        workbook = openpyxl.load_workbook("C:\\Users\\ysdong\\Desktop\\数学\\应用统计与R语言\\大作业\\doubanb.xlsx")
        sheets = workbook.get_sheet_names()
        generalsheet = workbook.get_sheet_by_name(sheets[2])

        detailsheet = workbook.get_sheet_by_name(sheets[5])
        generalId = generalsheet.cell(row=2, column=1).value
        detailId = detailsheet.cell(row=2, column=1).value

        generali=2
        detaili=2
        while generalId!=None and detailId!=None:
            plotfactor=0
            humourfactor=0
            actfactor=0
            while detailId == generalId:
                #calculate

                content=detailsheet.cell(row=detaili,column=2).value
                plotnum = 0
                humournum = 0
                actnum = 0
                while content!=None:

                    firstkeyword=0
                    firstSymbol=len(content)

                    tempPos = content.find("剧情")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 1

                    tempPos = content.find("推理")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 1

                    tempPos = content.find("故事")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 1

                    tempPos = content.find("情节")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 1

                    tempPos = content.find("强行")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 1

                    tempPos = content.find("结尾")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 1

                    tempPos = content.find("案件")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 1

                    tempPos = content.find("凶手")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 1

                    tempPos = content.find("侦探")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 1

                    tempPos = content.find("反转")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 1

                    tempPos = content.find("结局")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 1

                    tempPos = content.find("逻辑")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 1

                    tempPos = content.find("破案")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 1

                    tempPos = content.find("开头")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 1

                    tempPos = content.find("莫名其妙")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 1

                    tempPos = content.find("乱七八糟")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 1

                    tempPos = content.find("升华")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 1

                    tempPos = content.find("紧凑")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 1

                    tempPos = content.find("套路")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 1

                    tempPos = content.find("出乎意料")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 1

                    tempPos = content.find("出乎意料")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 2

                    tempPos = content.find("搞笑")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 2

                    tempPos = content.find("喜剧")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 2

                    tempPos = content.find("好笑")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 2

                    tempPos = content.find("演员")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 3

                    tempPos = content.find("陈思")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 3

                    tempPos = content.find("演技")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 3

                    tempPos = content.find("王宝强")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 3

                    tempPos = content.find("秦风")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 3

                    tempPos = content.find("刘昊然")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 3

                    tempPos = content.find("刘德华")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 3

                    tempPos = content.find("表演")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 3

                    tempPos = content.find("妹妹")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 3

                    tempPos = content.find("佟丽娅")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 3

                    tempPos = content.find("张子枫")
                    if tempPos < firstSymbol and tempPos != -1:
                        firstSymbol = tempPos
                        firstkeyword = 3


                    if firstkeyword==1:
                        plotnum=plotnum+1
                    elif firstkeyword==2:
                        humournum=humournum+1
                    elif firstkeyword==3:
                        actnum=actnum+1
                    else:
                        break

                    content = content[firstSymbol + 3:]
                    print(content)

                total=plotnum + humournum + actnum
                if total>0:
                    plotnum = plotnum / total
                    humournum = humournum / total
                    actnum = actnum / total
                else:
                    plotnum=0
                    humournum=0
                    actnum=0

                emotionscore=detailsheet.cell(row=detaili,column=7).value
                plotfactor = plotfactor + plotnum * emotionscore
                humourfactor = humourfactor + humournum * emotionscore
                actfactor = actfactor + actnum * emotionscore

                detaili=detaili+1
                detailId = detailsheet.cell(row=detaili, column=1).value


            generalsheet.cell(generali, 8, plotfactor)
            generalsheet.cell(generali, 9, humourfactor)
            generalsheet.cell(generali, 10, actfactor)

            generali = generali + 1
            generalId = generalsheet.cell(row=generali, column=1).value

        workbook.save("C:\\Users\\ysdong\\Desktop\\数学\\应用统计与R语言\\大作业\\doubanb.xlsx")
        workbook.close()





