# -*- coding: utf-8 -*-
# !/usr/bin/env python
import time
import urllib.request
import json
import pandas as pd

# client_id 为官网获取的AK， client_secret 为官网获取的SK
from openpyxl import load_workbook

client_id = '2011F6x7vE49iinihiDEGFhU'
client_secret = 'da5YIECSkvmjDbfOrzQRuGKyVd3IjLjp'


# 获取token
def get_token():
    host = 'https://aip.baidubce.com/oauth/2.0/token?grant_type=client_credentials&client_id=' + client_id + '&client_secret=' + client_secret
    request = urllib.request.Request(host)
    request.add_header('Content-Type', 'application/json; charset=UTF-8')
    response = urllib.request.urlopen(request)
    token_content = response.read()
    if token_content:
        token_info = json.loads(token_content)
        token_key = token_info['access_token']
    return token_key


def get_classify(content):
    print(content)
    token = get_token()
    url = 'https://aip.baidubce.com/rpc/2.0/nlp/v1/sentiment_classify'
    params = dict()
    params['text'] = content
    params = json.dumps(params).encode('utf-8')
    access_token = token
    url = url + "?access_token=" + access_token
    request = urllib.request.Request(url=url, data=params)
    request.add_header('Content-Type', 'application/json')
    response = urllib.request.urlopen(request)
    content = response.read()
    if content:
        try:content = content.decode('gb2312')
        except UnicodeDecodeError:
            return 'error'
        # print (content)
        data = json.loads(content)
        data = data['items'][0]
        sentiment = data['sentiment']
        return data['positive_prob'] * 10

def writeInto(emoji, star, text):
    tb_resource = load_workbook('求求.xlsx')
    tb_1 = tb_resource['Sheet1']
    loc = tb_1.max_row + 1
    tb_1[''.join(['A', str(loc)])] = emoji
    tb_1[''.join(['B', str(loc)])] = star
    tb_1[''.join(['C', str(loc)])] = text
    tb_resource.save('求求.xlsx')

if __name__ == '__main__':
    df = pd.read_excel('情感分析表.xlsx', '唐探1', usecols=[1, 2],
                       names=['评论', '评分'])
    list0 = df['评分'].values.tolist()
    comment = df['评论'].values.tolist()
    index_s = 55
    for index in range(index_s, len(comment)+1):
        text = comment[index]
        time.sleep(0.1)
        emoji = get_classify(text)
        if(emoji == 'error'):
            continue
        writeInto(emoji, list0[index], text)
